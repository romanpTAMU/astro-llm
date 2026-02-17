#!/usr/bin/env python3
"""
Build a Donny-style performance workbook from daily runs.

- Uses runs in data/runs/ (or --runs-dir).
- Rebalance twice per month: 2nd and 16th. Two periods per month:
  - Period 1: run closest to 2nd of M → run closest to 16th of M.
  - Period 2: run closest to 16th of M → run closest to 2nd of M+1.
- Loads portfolio from the start run; fetches daily equity performance via FMP API
  for each holding from start to end; writes cumulative return series and period return.
- Output: one sheet per rebalance period (e.g. "January 2026 (2nd-16th)", "January 2026 (16th-2nd)")
  + summary sheet "Summary" with daily differences by period, IR, and KP-style stats.

Requires: openpyxl, and FMP_API_KEY in .env for historical daily prices.
"""

from __future__ import annotations

import json
import re
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import typer

# Project root and path for agent imports
import sys
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from dotenv import load_dotenv
load_dotenv(PROJECT_ROOT / ".env")

from openpyxl import Workbook

from agent.config import load_config
from agent.data_apis import fetch_historical_daily_series
from agent.run_manager import find_all_portfolios


def parse_run_date(folder_name: str) -> Optional[date]:
    """Parse YYYY-MM-DD from run folder name YYYY-MM-DD_HH-MM-SS."""
    m = re.match(r"(\d{4}-\d{2}-\d{2})_\d{2}-\d{2}-\d{2}", folder_name)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), "%Y-%m-%d").date()
    except ValueError:
        return None


def find_run_closest_to_day(runs_with_dates: list[tuple[Path, date]], target_day: int, month: date) -> Optional[tuple[Path, date]]:
    """From list of (run_folder, run_date), find run in that month closest to target_day (e.g. 2 or 16)."""
    candidates = [(p, d) for p, d in runs_with_dates if d.year == month.year and d.month == month.month]
    if not candidates:
        return None
    # Prefer exact day, then closest
    def dist(item: tuple[Path, date]) -> int:
        d = item[1]
        return abs(d.day - target_day)
    return min(candidates, key=dist)


def get_rebalance_runs(runs_dir: Path) -> list[tuple[Path, date]]:
    """List all run folders that contain portfolio.json, with their run date. Sorted by date ascending."""
    if not runs_dir.exists():
        return []
    out = []
    for folder in runs_dir.iterdir():
        if not folder.is_dir():
            continue
        if not (folder / "portfolio.json").exists():
            continue
        d = parse_run_date(folder.name)
        if d is None:
            continue
        out.append((folder, d))
    out.sort(key=lambda x: x[1])
    return out


def load_portfolio(run_folder: Path) -> list[dict]:
    """Load holdings from run_folder/portfolio.json. Returns list of holding dicts (ticker, weight, ...)."""
    p = run_folder / "portfolio.json"
    if not p.exists():
        return []
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        return data.get("holdings", [])
    except Exception:
        return []


def build_monthly_periods(runs_with_dates: list[tuple[Path, date]]) -> list[tuple[str, Path, Path, date, date]]:
    """
    Build two rebalance periods per month (2nd and 16th):
    - (label, start_run_folder, end_run_folder, period_start_date, period_end_date).
    Period 1: run closest to 2nd of M → run closest to 16th of M.
    Period 2: run closest to 16th of M → run closest to 2nd of M+1.
    """
    if len(runs_with_dates) < 2:
        return []
    periods = []
    month_keys = sorted(set((d.year, d.month) for _, d in runs_with_dates))
    for year, month in month_keys:
        month_date = date(year, month, 1)
        month_label = month_date.strftime("%B %Y")  # e.g. January 2026
        if month == 12:
            next_year, next_month = year + 1, 1
        else:
            next_year, next_month = year, month + 1
        next_month_date = date(next_year, next_month, 1)

        run_2nd = find_run_closest_to_day(runs_with_dates, 2, month_date)
        run_16th = find_run_closest_to_day(runs_with_dates, 16, month_date)
        run_next_2nd = find_run_closest_to_day(runs_with_dates, 2, next_month_date)

        # Period 1: 2nd → 16th (same month)
        if run_2nd is not None and run_16th is not None:
            start_folder, start_date = run_2nd
            end_folder, end_date = run_16th
            if start_date < end_date:
                periods.append((f"{month_label} (2nd-16th)", start_folder, end_folder, start_date, end_date))

        # Period 2: 16th → 2nd (next month)
        if run_16th is not None and run_next_2nd is not None:
            start_folder, start_date = run_16th
            end_folder, end_date = run_next_2nd
            if start_date < end_date:
                periods.append((f"{month_label} (16th-2nd)", start_folder, end_folder, start_date, end_date))
    return periods


def fetch_daily_returns_for_holdings(
    holdings: list[dict],
    period_start: date,
    period_end: date,
    fmp_key: str,
    rate_limit_delay: float = 0.25,
) -> dict[str, list[tuple[date, float]]]:
    """
    For each holding (ticker), fetch daily close series from FMP and return
    ticker -> [(date, cumulative_return)] where cumulative_return = close_t / close_first.
    """
    tickers = [h["ticker"] for h in holdings if h.get("ticker")]
    result = {}
    for ticker in tickers:
        series = fetch_historical_daily_series(ticker, period_start, period_end, fmp_key)
        if not series:
            result[ticker] = []
            continue
        first_close = series[0][1]
        if first_close == 0:
            result[ticker] = []
            continue
        cum = [(d, close / first_close) for d, close in series]
        result[ticker] = cum
        time.sleep(rate_limit_delay)
    return result


def period_return_from_series(series: list[tuple[date, float]]) -> Optional[float]:
    """Period return = last cumulative value (series is already close/first_close)."""
    if not series:
        return None
    return series[-1][1]


def write_month_sheet(
    wb: Workbook,
    sheet_name: str,
    holdings: list[dict],
    daily_series: dict[str, list[tuple[date, float]]],
    period_start: date,
    period_end: date,
    port_cum: list[tuple[date, float]],
    sp500_series: list[tuple[date, float]],
    beta: Optional[float],
    alpha: Optional[float],
) -> None:
    """Write one sheet: Symbol, Exchange, Sector, Ticker, Weight, Return, period_start, period_end, then daily cumulative cols.
    After holdings: Total Return row (weighted port), S&P 500 row, Beta row, Alpha row."""
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
    all_dates = set()
    for series in daily_series.values():
        for d, _ in series:
            all_dates.add(d)
    sorted_dates = sorted(all_dates) if all_dates else []
    # Header: 1=Symbol, 2=Exchange, 3=Sector, 4=Ticker, 5=Weight, 6=Return, 7=period_start, 8=period_end, 9+=dates
    ws.cell(1, 5, "Weight")
    ws.cell(1, 6, "Return")
    ws.cell(1, 7, period_start)
    ws.cell(1, 8, period_end)
    for c, d in enumerate(sorted_dates, start=9):
        ws.cell(1, c, d)
    # Holdings rows
    for row_idx, h in enumerate(holdings, start=2):
        ticker = h.get("ticker", "")
        weight = h.get("weight", 0)
        sector = h.get("sector", "")
        ws.cell(row_idx, 1, ticker)
        ws.cell(row_idx, 2, "")
        ws.cell(row_idx, 3, sector)
        ws.cell(row_idx, 4, ticker)
        ws.cell(row_idx, 5, weight)
        series = daily_series.get(ticker, [])
        ret = period_return_from_series(series)
        if ret is not None:
            ws.cell(row_idx, 6, ret)
        ws.cell(row_idx, 7, period_start)
        ws.cell(row_idx, 8, period_end)
        date_to_cum = {d: v for d, v in series}
        for c, d in enumerate(sorted_dates, start=9):
            val = date_to_cum.get(d)
            if val is not None:
                ws.cell(row_idx, c, val)
    # Total Return row (weighted portfolio)
    next_row = len(holdings) + 2
    ws.cell(next_row, 1, "Total Return")
    ws.cell(next_row, 6, period_return_from_series(port_cum))
    port_date_to_cum = dict(port_cum)
    for c, d in enumerate(sorted_dates, start=9):
        val = port_date_to_cum.get(d)
        if val is not None:
            ws.cell(next_row, c, val)
    # S&P 500 row
    next_row += 1
    ws.cell(next_row, 1, "S&P 500")
    ws.cell(next_row, 6, period_return_from_series(sp500_series))
    sp_date_to_cum = dict(sp500_series)
    for c, d in enumerate(sorted_dates, start=9):
        val = sp_date_to_cum.get(d)
        if val is not None:
            ws.cell(next_row, c, val)
    # Beta and Alpha rows
    next_row += 1
    ws.cell(next_row, 1, "Beta")
    if beta is not None:
        ws.cell(next_row, 6, beta)
    next_row += 1
    ws.cell(next_row, 1, "Alpha")
    if alpha is not None:
        ws.cell(next_row, 6, alpha)


def portfolio_daily_returns(
    holdings: list[dict],
    daily_series: dict[str, list[tuple[date, float]]],
) -> list[tuple[date, float]]:
    """Compute portfolio-weighted daily cumulative return series (same dates across stocks)."""
    if not holdings or not daily_series:
        return []
    # All dates
    all_dates = set()
    for series in daily_series.values():
        for d, _ in series:
            all_dates.add(d)
    sorted_dates = sorted(all_dates)
    if not sorted_dates:
        return []
    weight_by_ticker = {h["ticker"]: h.get("weight", 0) for h in holdings}
    total_weight = sum(weight_by_ticker.values()) or 1
    result = []
    for d in sorted_dates:
        cum = 0.0
        for ticker, series in daily_series.items():
            date_to_cum = dict(series)
            if d in date_to_cum:
                w = weight_by_ticker.get(ticker, 0) / total_weight
                cum += w * date_to_cum[d]
        result.append((d, cum))
    return result


def daily_differences(portfolio_daily_cum: list[tuple[date, float]]) -> list[float]:
    """Convert cumulative series to daily period-over-period differences (daily returns)."""
    if len(portfolio_daily_cum) < 2:
        return []
    diffs = []
    for i in range(1, len(portfolio_daily_cum)):
        prev = portfolio_daily_cum[i - 1][1]
        curr = portfolio_daily_cum[i][1]
        if prev != 0:
            diffs.append((curr / prev) - 1.0)
        else:
            diffs.append(0.0)
    return diffs


# S&P 500 proxy ticker for benchmark
SP500_TICKER = "SPY"


def fetch_sp500_series(
    period_start: date,
    period_end: date,
    fmp_key: str,
) -> list[tuple[date, float]]:
    """Fetch S&P 500 (SPY) daily close series; return [(date, cumulative_return)] with first = 1."""
    series = fetch_historical_daily_series(SP500_TICKER, period_start, period_end, fmp_key)
    if not series:
        return []
    first_close = series[0][1]
    if first_close == 0:
        return []
    return [(d, close / first_close) for d, close in series]


def aligned_daily_returns(
    port_cum: list[tuple[date, float]],
    market_cum: list[tuple[date, float]],
) -> tuple[list[float], list[float]]:
    """Return (port_daily_returns, market_daily_returns) for dates present in both series."""
    port_by_date = dict(port_cum)
    market_by_date = dict(market_cum)
    common_dates = sorted(set(port_by_date) & set(market_by_date))
    if len(common_dates) < 2:
        return [], []
    port_rets = []
    market_rets = []
    for i in range(1, len(common_dates)):
        d_prev, d_curr = common_dates[i - 1], common_dates[i]
        p_prev, p_curr = port_by_date[d_prev], port_by_date[d_curr]
        m_prev, m_curr = market_by_date[d_prev], market_by_date[d_curr]
        port_rets.append((p_curr / p_prev) - 1.0 if p_prev else 0.0)
        market_rets.append((m_curr / m_prev) - 1.0 if m_prev else 0.0)
    return port_rets, market_rets


def compute_beta_alpha(
    port_cum: list[tuple[date, float]],
    market_cum: list[tuple[date, float]],
) -> tuple[Optional[float], Optional[float]]:
    """
    Beta = Cov(port daily ret, market daily ret) / Var(market daily ret).
    Alpha = port_period_return - beta * market_period_return.
    Returns (beta, alpha).
    """
    port_rets, market_rets = aligned_daily_returns(port_cum, market_cum)
    if len(port_rets) < 2:
        return None, None
    port_period = port_cum[-1][1] if port_cum else 1.0
    market_period = market_cum[-1][1] if market_cum else 1.0
    n = len(port_rets)
    mean_p = sum(port_rets) / n
    mean_m = sum(market_rets) / n
    cov = sum((p - mean_p) * (m - mean_m) for p, m in zip(port_rets, market_rets)) / (n - 1)
    var_m = sum((m - mean_m) ** 2 for m in market_rets) / (n - 1)
    if var_m == 0:
        return None, None
    beta = cov / var_m
    alpha = (port_period - 1.0) - beta * (market_period - 1.0)  # period returns as (cum - 1)
    return beta, alpha


def chain_cumulative_series(
    period_series_list: list[list[tuple[date, float]]],
) -> list[tuple[date, float]]:
    """Chain per-period cumulative series (each starts at 1) into one series. Scale each period by product of prior period returns."""
    chained = []
    scale = 1.0
    for series in period_series_list:
        if not series:
            continue
        for d, v in series:
            chained.append((d, scale * v))
        scale *= series[-1][1]
    return chained


def combined_total_return_beta_alpha(
    summary_rows: list[tuple[str, list[float], list[tuple[date, float]], list[tuple[date, float]]]],
) -> tuple[Optional[float], Optional[float], Optional[float], Optional[float]]:
    """
    Combine all periods: chain port and market cumulative series, then compute:
    - Total return (multiplicative): product of period returns = chained_port[-1].
    - Total market return: chained_market[-1].
    - Beta from full chained daily returns.
    - Alpha = (total_port - 1) - beta * (total_market - 1).
    Returns (total_return_cum, beta, alpha, total_market_cum).
    """
    if not summary_rows:
        return None, None, None, None
    port_series_list = [r[2] for r in summary_rows]
    market_series_list = [r[3] for r in summary_rows]
    chained_port = chain_cumulative_series(port_series_list)
    chained_market = chain_cumulative_series(market_series_list)
    if not chained_port or not chained_market:
        return None, None, None, None
    total_return = chained_port[-1][1]
    total_market = chained_market[-1][1]
    beta, alpha = compute_beta_alpha(chained_port, chained_market)
    if beta is not None and alpha is None:
        alpha = (total_return - 1.0) - beta * (total_market - 1.0)
    return total_return, beta, alpha, total_market


def main(
    runs_dir: Path = typer.Option(Path("data/runs"), "--runs-dir", help="Base directory of daily runs"),
    out: Path = typer.Option(Path("data/daily_performance_report.xlsx"), "--out", help="Output Excel path"),
    from_month: Optional[str] = typer.Option(None, "--from-month", help="Start month YYYY-MM (default: first run month)"),
    to_month: Optional[str] = typer.Option(None, "--to-month", help="End month YYYY-MM (default: last run month)"),
    rate_limit: float = typer.Option(0.25, "--rate-limit", help="Seconds between FMP API calls per ticker"),
) -> None:
    runs_dir = Path(runs_dir)
    runs_base = PROJECT_ROOT / runs_dir if not runs_dir.is_absolute() else runs_dir
    cfg = load_config()
    fmp_key = cfg.fmp_api_key
    if not fmp_key:
        typer.echo("[ERROR] FMP_API_KEY required for historical daily prices. Set in .env")
        raise typer.Exit(1)

    runs_with_dates = get_rebalance_runs(runs_base)
    if not runs_with_dates:
        typer.echo(f"[ERROR] No runs with portfolio.json in {runs_base}")
        raise typer.Exit(1)

    periods = build_monthly_periods(runs_with_dates)
    if from_month:
        try:
            y, m = map(int, from_month.split("-"))
            filter_start = date(y, m, 1)
            periods = [(lb, a, b, s, e) for lb, a, b, s, e in periods if s >= filter_start]
        except ValueError:
            pass
    if to_month:
        try:
            y, m = map(int, to_month.split("-"))
            filter_end = date(y, m, 28)
            periods = [(lb, a, b, s, e) for lb, a, b, s, e in periods if s <= filter_end]
        except ValueError:
            pass

    if not periods:
        typer.echo("[ERROR] No periods to report (need runs closest to 2nd and 16th)")
        raise typer.Exit(1)

    wb = Workbook()
    default_sheet = wb.active
    summary_rows = []  # for Summary sheet: (month_name, daily_diffs list), then IR row, KP stats

    for label, start_folder, end_folder, period_start, period_end in periods:
        typer.echo(f"Period {label}: {period_start} -> {period_end} (run {start_folder.name})")
        holdings = load_portfolio(start_folder)
        if not holdings:
            typer.echo(f"  [WARN] No holdings in {start_folder}")
            continue
        daily_series = fetch_daily_returns_for_holdings(
            holdings, period_start, period_end, fmp_key, rate_limit_delay=rate_limit
        )
        port_cum = portfolio_daily_returns(holdings, daily_series)
        sp500_series = fetch_sp500_series(period_start, period_end, fmp_key)
        time.sleep(rate_limit)  # rate limit after SPY fetch
        beta, alpha = compute_beta_alpha(port_cum, sp500_series)
        write_month_sheet(
            wb, label, holdings, daily_series, period_start, period_end,
            port_cum, sp500_series, beta, alpha,
        )
        diffs = daily_differences(port_cum)
        summary_rows.append((label, diffs, port_cum, sp500_series))

    # Summary sheet
    ws_sum = wb.create_sheet(title="Summary", index=0)
    ws_sum.cell(1, 2, "Daily Difference")
    max_days = max(len(r[1]) for r in summary_rows) if summary_rows else 0
    for col in range(max_days):
        ws_sum.cell(2, col + 2, col)
    for row_idx, row in enumerate(summary_rows, start=3):
        label, diffs = row[0], row[1]
        ws_sum.cell(row_idx, 1, label)
        for col, v in enumerate(diffs):
            ws_sum.cell(row_idx, col + 2, v)
    # Combined (all periods): Total Return, S&P 500 Return, Beta, Alpha
    total_ret, combined_beta, combined_alpha, market_ret = combined_total_return_beta_alpha(summary_rows)
    base_row = len(summary_rows) + 3
    ws_sum.cell(base_row, 1, "COMBINED (all periods)")
    base_row += 1
    ws_sum.cell(base_row, 1, "Total Return (cumulative)")
    if total_ret is not None:
        ws_sum.cell(base_row, 2, total_ret)
        ws_sum.cell(base_row, 3, f"{(total_ret - 1) * 100:+.2f}%")
    base_row += 1
    ws_sum.cell(base_row, 1, "S&P 500 Return (cumulative)")
    if market_ret is not None:
        ws_sum.cell(base_row, 2, market_ret)
        ws_sum.cell(base_row, 3, f"{(market_ret - 1) * 100:+.2f}%")
    base_row += 1
    ws_sum.cell(base_row, 1, "Beta")
    if combined_beta is not None:
        ws_sum.cell(base_row, 2, combined_beta)
    base_row += 1
    ws_sum.cell(base_row, 1, "Alpha")
    if combined_alpha is not None:
        ws_sum.cell(base_row, 2, combined_alpha)
        ws_sum.cell(base_row, 3, f"{combined_alpha * 100:+.2f}%")
    base_row += 2
    # IR: annualized (mean daily return / std daily return) * sqrt(252).
    all_diffs = []
    for r in summary_rows:
        all_diffs.extend(r[1])
    if all_diffs:
        import math
        mean_d = sum(all_diffs) / len(all_diffs)
        var = sum((x - mean_d) ** 2 for x in all_diffs) / len(all_diffs) if len(all_diffs) > 1 else 0
        std_d = math.sqrt(var) if var > 0 else 0
        ir = (mean_d / std_d * math.sqrt(252)) if std_d != 0 else 0
        ws_sum.cell(base_row, 1, "IR")
        ws_sum.cell(base_row, 2, ir)
        base_row += 1
    # KP-style: total trading days, fraction positive
    total_days = len(all_diffs)
    positive_days = sum(1 for x in all_diffs if x > 0)
    frac = positive_days / total_days if total_days else 0
    ws_sum.cell(base_row + 1, 1, "KP STATS")
    ws_sum.cell(base_row + 1, 2, total_days)
    ws_sum.cell(base_row + 1, 3, frac)
    ws_sum.cell(base_row + 2, 2, positive_days)
    ws_sum.cell(base_row + 2, 3, frac)

    wb.remove(default_sheet)

    out = Path(out)
    out_path = PROJECT_ROOT / out if not out.is_absolute() else out
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    typer.echo(f"[SUCCESS] Wrote {out_path}")


if __name__ == "__main__":
    typer.run(main)
