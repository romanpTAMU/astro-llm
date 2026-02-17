"""Quick read of COMBINED (all periods) from data/daily_performance_report.xlsx Summary sheet."""
from pathlib import Path
from openpyxl import load_workbook

path = Path(__file__).parent.parent / "data" / "daily_performance_report.xlsx"
if not path.exists():
    print(f"Report not found: {path}")
    exit(1)
wb = load_workbook(path, read_only=True, data_only=True)
ws = wb["Summary"]
print("COMBINED (all periods)")
print("-" * 50)
for row in ws.iter_rows(min_row=1, max_row=25, min_col=1, max_col=4, values_only=True):
    a = row[0]
    b = row[1] if len(row) > 1 else None
    c = row[2] if len(row) > 2 else None
    if a in ("Total Return (cumulative)", "S&P 500 Return (cumulative)", "Beta", "Alpha"):
        print(f"  {a}: {b}  {c or ''}")
wb.close()
